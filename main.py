import yfinance as yf
from openpyxl import load_workbook
import logging
import json
import threading
import time
import schedule
import os
import sys

'''
	Get stats from yahoo finance 
	for each ticker symbol in the spreadsheet
'''
def get_stats(ticker, row_count):
    info = yf.Tickers(ticker).tickers[ticker].info
    update_stock_price(info['currentPrice'], row_count)


'''
	Access and modify the cell values of column B
	with new stock prices
'''
def update_stock_price(price, row_count):
    sheet.cell(row=row_count, column=2).value = price
    workbook.save(sheetPath)  # Save the changes to the spreadsheet


def main():
    global workbook, sheet, sheetPath, config_path
    threads = []

    # Load the config file and spreadsheet
    try:
        with open(config_path) as f:
            config = json.load(f)
            sheetPath = config["path"]

    except FileNotFoundError:
        logging.error("Config file not found")
        return

    try:
        workbook = load_workbook(sheetPath)
        sheet = workbook.active

    except FileNotFoundError:
        logging.error("Excel file not found")
        return

    '''
		Go through each ticker symbol in column A
		and get the stats for that symbol
	'''
    row_count = 2
    while True:
        # Read the cell values of column A
        ticker = sheet.cell(row=row_count, column=1).value
        if ticker is None:
            break

        # Create a new thread for each ticker symbol
        thread = threading.Thread(target=get_stats, args=[ticker, row_count])
        thread.start()
        threads.append(thread)
        row_count += 1

    # Wait for all threads to finish
    for thread in threads:
        thread.join()

    date_time = time.strftime("%d/%m/%Y %H:%M:%S")
    logging.info(f"Updated {date_time}")  # Log the time of the update


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)

    # Determine if the application is a .exe
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    # Or a script file
    elif __file__:
        application_path = os.path.dirname(__file__)

    # Get the path to the config file
    config_name = 'StonksConfig.json'
    config_path = os.path.join(application_path, config_name)

    try:
        # Load config.json and get interval value
        with open(config_path) as f:
            config = json.load(f)
            interval = config["interval"]

    except FileNotFoundError:
        logging.critical("Config file not found")
        sys.exit(1)

    # Run every "interval" minutes
    main()
    schedule.every(interval).minutes.do(main)
    while True:
        schedule.run_pending()
        time.sleep(1)
